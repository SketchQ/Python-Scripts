#! pyhton3
# multiplicationTable.py - Creates an Excel spreadsheet table from command line

# This progrom will do :
    # 1. Take N from command line
    # 2. Will create an NxN multiplication table in excel

# Code needs to do :
    # 1.Take an argument from command line
    # 2.Create a table with that argument


import sys,openpyxl
from openpyxl.cell import cell
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title ='Multiplication Table'
wb.save('multiplicationTable.xlsx')
prompt =''.join(sys.argv[1:])
'''prompt = input()
prompt =''.join(prompt)'''


if prompt.isnumeric():
    for i in range(int(prompt)+1):
        for y in range(int(prompt)+1):
            isHeader = False

            if y == 0 and i == 0:
                isHeader = True
                n = ''
            
            elif i == 0:
                isHeader=True
                n = y
            
            elif y == 0:
                isHeader = True
                n = i

            else:
                n = i*y

            c = sheet.cell(row=y+1,column=i+1)

            if isHeader:
                c.font = Font(bold=True)

            c.value = n
    f= 'multiplication_table_' +str(prompt) + '.xlsx'
    
    wb.save(f)

    print('Saved as ' + f)

    sheet.freeze_panes = 'B2'  
    wb.save(f)   
else:
    print('You need to enter a number')

'''
        sheet.cell(row=i+1,column=1).value = i
        sheet.cell(row=1,column=i+1).value = i
        sheet.cell(row=i+1,column=1).font = Font(bold=True,size=14)
        sheet.cell(row=1,column=i+1).font = Font(bold=True,size=14)
        sheet.cell(row=i+1,column=2).value = '=A2:A26*$B$1'
        sheet.cell(row=i+1,column=3).value = '=A2:A26*$C$1'
        sheet.cell(row=i+1,column=4).value = '=A2:A26*$D$1'
        sheet.cell(row=i+1,column=5).value = '=A2:A26*$E$1'
        sheet.cell(row=i+1,column=6).value = '=A2:A26*$F$1'
        sheet.cell(row=i+1,column=7).value = '=A2:A26*$G$1'
'''